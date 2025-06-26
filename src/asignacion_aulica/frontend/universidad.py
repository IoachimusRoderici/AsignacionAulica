import pandas as pd
from datetime import time
import re

from asignacion_aulica.frontend.excepciones_universidad import *
from asignacion_aulica.get_asset_path import get_asset_path
from asignacion_aulica.frontend import funciones_de_traduccion
from asignacion_aulica.backend.lógica_de_asignación   import asignar

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


class Universidad:
    def __init__(self, 
        df_list = pd.read_excel(get_asset_path('edificios_default/Universidad.xlsx'), sheet_name=None)
        #df_list = pd.read_excel("./src/assets/edificios_default/Universidad.xlsx", sheet_name=None)
    ):        
        
        self.edificios = df_list['Edificios']
        self.aulas = df_list['Aulas']
        self.carreras = df_list['Carreras']
        self.materias = df_list['Materias']
        self.horarios = df_list['Clases'].fillna("")
        self.horarios = self.horarios.head(2)  



    # Sector de edificios:

    def columnas_edificios(self): 
        """
        Metodo que retorna lista de nombres de columnas, del dataframe de Edificios

        Parameters
        -None
        Returns
            La lista de columnas de los edificios
        TYPE
            List[str]
        """
        return self.edificios.columns.tolist()
    def mostrar_edificios(self):
        """
        Retorna el dataframe de los edificios instanciados

        Parameters
        -None
        Returns
            La lista de columnas de los edificios
        TYPE
            DataFrame
        """
        return self.edificios
    def nombres_edificios(self):
        """
        Metodo para mostrar los nombres de los edificios instanciados.
        Sirve para el menu dropdown al intentar crear aulas.

        Parameters
        -None
        Returns
            La lista de nombres de los edificios instanciados
        TYPE
            List[str]
        """
        return self.edificios.iloc[:,0].tolist()
    def agregar_edificio(self, nombre_edificio:str):
        """
        Metodo para agregar un edificio a la universidad.

        Parameters
        ----------
        nombre_edificio : str
            El nombre del edificio a agregar. Funciona como clave primaria.
        Returns
            None
        Throws:
            ElementoYaExistenteException , si se trata de agregar un edificio ya existente.
        """
        nombre_edificio = nombre_edificio.strip()
        if nombre_edificio == "":
            raise(ElementoInvalidoException("No se puede agregar un edificio sin nombre"))

        if nombre_edificio in self.nombres_edificios():
            raise(ElementoDuplicadoException("Ya existe un edificio con ese nombre"))

        aux_dict = {'Edificio':nombre_edificio}
        for col in ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']:
            aux_dict[col] = "9:00-23:00"
        aux_dict['Domingo'] = "CERRADO"
        aux_row = pd.DataFrame([aux_dict])
        self.edificios = pd.concat([self.edificios, aux_row], ignore_index=True)
    def eliminar_edificio(self, nombre_edificio:str):
        """
        Metodo para eliminar un edificio existente de la universidad

        Parameters
        ----------
        nombre_edificio : str
            El nombre del edificio a eliminar. Funciona como clave primaria.
        Returns
            None
        Throws:
            ElementoNoExisteException, si se trata de borrar un edificio que no existe en sistema.
            ElementoTieneDependenciasException , si se trata de agregar un edificio ya existente.
        """
        if nombre_edificio in self.aulas['Edificio'].values:
            raise(ElementoTieneDependenciasException("Hay aulas que contienen este edificio, eliminacion cancelada."))
        if nombre_edificio not in self.nombres_edificios():
            raise(ElementoNoExisteException("El edificio que desea borrar no existe en el sistema."))
        
        self.edificios = self.edificios[self.edificios['Edificio'] != nombre_edificio].reset_index(drop=True)    
    def modificar_edificio(self, nombre_edificio:str, columna_a_modificar:str, valor_nuevo:str):
        """
        Modifica el valor de una columna específica para el edificio dado.

        Parameters
        ----------
        nombre_edificio : str
            Nombre del edificio a modificar (clave primaria, debe estar en la primera columna).
        columna_a_modificar : str
            Nombre de la columna a modificar.
        valor_nuevo : str
            Valor nuevo a establecer en la celda correspondiente.
        
        Returns
        -------
        None
        Throws
            - ElementoInvalidoException , si se trata de ingresar algun parametro vacio.

        """
        if nombre_edificio=="":
            raise(ElementoInvalidoException("Debe ingresar un edificio a modificar."))
        if columna_a_modificar=="":
            raise(ElementoInvalidoException("Debe elegir el dia al que quiera modificar su horario."))
        if columna_a_modificar not in self.columnas_edificios():
            raise(ElementoInvalidoException(f"La columna que desea modificar ({columna_a_modificar}) no se encuentra entre los datos del edificio ({self.columnas_edificios()}.)"))
        if valor_nuevo=="":
            raise(ElementoInvalidoException("No se puede ingresar una cadena vacia como valor nuevo."))

        self.edificios.at[self.indice_edificio(nombre_edificio), columna_a_modificar] = valor_nuevo
    def modificar_horario_edificio(self, nombre_edificio:str, dia:str, hora1:int, hora2:int, minuto1:int, minuto2:int):

        if (
            hora1 not in range(0,24) or
            hora2 not in range(0,24) or
            minuto1 not in range(0,60) or
            minuto2 not in range(0,60)
        ):
            raise(HorarioInvalidoException("Error: Los datos de horario deben estar en un rango de 0-23 horas y 0-59 minutos."))

        if time(hora1, minuto1) < time(hora2, minuto2):
            self.modificar_edificio(nombre_edificio, dia, f"{hora1}:{minuto1:02}-{hora2}:{minuto2:02}")
        else:
            raise(HorarioInvalidoException("La hora de cierre no puede ser menor que la de apertura"))
    def indice_edificio(self, nombre_edificio:str):
        """"Metodo auxiliar que retorna el indice de un edificio en el dataframe. Para evitar reusar codigo."""
        filtro = self.edificios[self.edificios['Edificio'] == nombre_edificio]
        return filtro.index[0]
    def valor_de_edificio(self, nombre_edificio:str, dia:str):
        """Metodo que retorna la cadena de horario de un edificio, sin importar si esta cerrado o no.
        Usado para parsear los horarios completos, o para instanciar horarios por defecto en aulas.
        """
        return self.edificios.at[self.indice_edificio(nombre_edificio), dia]
    def horario_segmentado_edificio(self, nombre_edificio:str, dia:str):
        """Metodo que retorna el horario de un edificio, pero segmentado en los cuatro valores componentes.
        Sigue el orden que tenia la cadena (Hora apertura, minuto apertura, hora cierre, minuto cierre).
        Si ese edificio estaba cerrado, retorna cadenas vacias."""
        cadena = self.valor_de_edificio(nombre_edificio, dia)
        matches = re.findall(r'\d+', cadena) 
        if len(matches)==4:
            return [match.zfill(2) for match in matches]
        else:
            raise(HorarioInvalidoException("Error, no se puede parsear el horario: " + cadena))


    



    # Sector de aulas
    def columnas_aulas(self):
        return self.aulas.columns.tolist()
    def mostrar_aulas(self):
        return self.aulas
    def nombres_aulas(self):
        """
        Metodo para mostrar los nombres de las aulas instanciadas.

        Parameters
        -None
        Returns
            La lista de nombres de las aulas instanciadas
        TYPE
            List[str]
        """
        return self.aulas.iloc[:,0].tolist()    
    def indice_aula(self, nombre_aula:str):
        """"Metodo auxiliar que retorna el indice de un aula en el dataframe. Para evitar reusar codigo."""
        filtro = self.aulas[self.aulas['Aula'] == nombre_aula]
        return filtro.index[0]
    def agregar_aula(self, identificador_aula:str , capacidad:str, edificio_aula:str):
        """Metodo para agregar un aula al dataframe de aulas. Verifica unicidad y edificio existente."""
        if edificio_aula not in self.nombres_edificios():
            raise(ElementoNoExisteException("No se agrego el aula con un edificio valido."))
        if identificador_aula in self.nombres_aulas():
            raise(ElementoDuplicadoException("Ya existe un aula con ese nombre en el sistema."))
        try:
            test_aux = int(capacidad)
            if test_aux<=0:
                raise Exception
        except Exception as e:
            raise(ElementoInvalidoException("La capacidad debe ser un numero entero positivo"))


        aux_dict = {col:None for col in self.columnas_aulas()}
        aux_dict['Aula'] = identificador_aula    #   Primer columna es identificador aula. Escribo
        aux_dict['Capacidad'] = capacidad
        aux_dict['Edificio'] = edificio_aula        #   Ultima columna es edificio. Escribo.
        for col in ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]:
            aux_dict[col] = self.valor_de_edificio(edificio_aula, col)
        aux_row = pd.DataFrame([aux_dict])
        self.aulas = pd.concat([self.aulas, aux_row], ignore_index=True)
    def eliminar_aula(self, nombre_aula): #TODO Anda, pero definir si requiere un check de dependencias antes de eliminar aulas.
        """
        Metodo para eliminar un edificio existente de la universidad

        Parameters
        ----------
        nombre_edificio : str
            El nombre del edificio a eliminar. Funciona como clave primaria.
        Returns
            None
        Throws:
            ElementoNoExisteException, si se trata de borrar un edificio que no existe en sistema.
        """
        # TODO posible chequeo de donde se usa el aula. Pensar, cuando esten las materias listas.
        #if nombre_aula in self.aulas['Edificio'].values:
        #    raise(ElementoTieneDependenciasException("Hay aulas que contienen este edificio, eliminacion cancelada."))
        if nombre_aula not in self.nombres_aulas():
            raise(ElementoNoExisteException("El aula que desea borrar no existe en el sistema."))
        
        self.aulas = self.aulas[self.aulas['Aula'] != nombre_aula].reset_index(drop=True)
    def modificar_aula(self, nombre_aula:str, columna_a_modificar:str, valor_nuevo:str):
        """
        Modifica el valor de una columna específica para el aula dada.

        Parameters
        ----------
        nombre_aula : str
            Nombre del aula a modificar (clave primaria, debe seguir la convencion de nombres dada en el README de formato).
        columna_a_modificar : str
            Nombre de la columna a modificar.
        valor_nuevo : str
            Valor nuevo a establecer en la celda correspondiente.
        Returns
        -------
        None
        Throws
            - ElementoInvalidoException , si se trata de ingresar algun parametro vacio.

        """
        if nombre_aula not in self.nombres_aulas():
            raise(ElementoInvalidoException("Para modificar un aula, debe elegir un aula que este en el sistema."))
        if columna_a_modificar not in self.columnas_aulas():
            raise(ElementoInvalidoException(f"La columna que desea modificar ({columna_a_modificar})" + 
                                            f"no se encuentra entre los datos del edificio ({self.columnas_aulas()}.)"))
        
        valor_nuevo = str(valor_nuevo)
        
        if valor_nuevo=="":
            raise(ElementoInvalidoException("No se puede ingresar una cadena vacia como valor nuevo."))
        
        self.aulas.at[self.indice_aula(nombre_aula), columna_a_modificar] = valor_nuevo
    def modificar_horario_aula(self, nombre_aula:str, dia:str, hora1:int, hora2:int, minuto1:int, minuto2:int):

        if (
            hora1 not in range(0,24) or
            hora2 not in range(0,24) or
            minuto1 not in range(0,60) or
            minuto2 not in range(0,60)
        ):
            raise(HorarioInvalidoException("Error: Los datos de horario deben estar en un rango de 0-23 horas y 0-59 minutos."))

        if time(hora1, minuto1) < time(hora2, minuto2):
            self.modificar_aula(nombre_aula, dia, f"{hora1}:{minuto1:02}-{hora2}:{minuto2:02}")
        else:
            raise(HorarioInvalidoException("La hora de cierre no puede ser menor que la de apertura"))
    def valor_de_aula(self, nombre_aula:str, dia:str):
        """Metodo que retorna la cadena de horario de un aula, sin importar si esta cerrado o no.
        """
        return self.aulas.at[self.indice_aula(nombre_aula), dia]
    def horario_segmentado_aula(self, nombre_aula:str, dia:str):
        """Metodo que retorna el horario de un aula, pero segmentado en los cuatro valores componentes.
        Sigue el orden que tenia la cadena (Hora apertura, minuto apertura, hora cierre, minuto cierre).
        Si ese edificio estaba cerrado, explota en horarioinvalidoexception"""
        cadena = self.valor_de_aula(nombre_aula, dia)
        matches = re.findall(r'\d+', cadena) 
        if len(matches)==4:
            return [match.zfill(2) for match in matches]
        else:
            raise(HorarioInvalidoException("Error, no se puede parsear el horario: " + cadena))
    def aulas_de_edificio(self, nombre_edificio:str):
        return self.aulas[self.aulas['Edificio'] == nombre_edificio]['Aula'].tolist()





################3
###############CARRERAS
    def mostrar_carreras(self):
        return self.carreras
    def nombres_carreras(self):
        return self.carreras.iloc[:,0].tolist()


############# Materias
    def mostrar_materias(self):
        return self.materias.fillna("")

    def nombres_materias(self):
        mostrable = self.materias.copy()
        return mostrable.iloc[:,0].tolist().drop_duplicates()
    
    #Usar cuando Materias este completamente implementado
    def nombres_materias_concatenados(self):    #TODO agregar parametro de validacion con carrera de la materia

        ret_df = pd.DataFrame()
        ret_df['Combinado'] = (
        self.materias['Código'].astype(str) + " - " +
        self.materias['Nombre'] + " - " +
        self.materias['Comisión'].astype(str)
        )
        
        return ret_df.iloc[:,0].tolist() 

#   clases

    def columnas_horarios(self):
        return self.horarios.columns.tolist()
    def nombres_horarios(self):
        return self.horarios.iloc[:,0].tolist()    
    def mostrar_horarios(self):
        return self.horarios
    
    def agregar_horario(self, nombre:str, dia:str, hora1, hora2, minuto1, minuto2, cantidad, edificio=""):
        """Metodo para agregar un horario al sistema.
        Dado que siempre se trata de modificar una materia existente, este metodo solo sera usado
        dentro de modificar_horario, cuando no se encuentra una combinacion de nombre y dia valida."""

        aux_dict = {col:"" for col in self.horarios.columns.tolist()}
        aux_dict['Nombre'] = nombre
        aux_dict['Día'] = dia
        aux_dict['Horario de inicio'] = f"{hora1}:{minuto1}"
        aux_dict['Horario de fin'] = f"{hora2}:{minuto2}"
        aux_dict['Cantidad de alumnos'] = int(cantidad)
        aux_dict['Edificio preferido'] = "Anasagasti 2" # Des-hardcodear despues
        # Despues implementar equipamiento

        aux_row = pd.DataFrame([aux_dict])
        self.horarios = pd.concat([self.horarios, aux_row], ignore_index=True)        

    def modificar_horario(self, nombre_horario:str, dia:str, hora1, hora2, minuto1, minuto2, cantidad=50, edificio=""):
        
        try:
            cantidad = int(cantidad)
        except Exception:
            raise(ValueError("La cantidad debe ser un numero positivo"))

        if nombre_horario not in self.nombres_horarios():
            raise(ElementoInvalidoException("Para modificar un horario, debe elegir un horario que este en el sistema."))
        if dia not in ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']:
            raise(ElementoInvalidoException(f"Se trato de agregar un horario a un dia no valido."))
        if (
            hora1 not in range(0,24) or
            hora2 not in range(0,24) or
            minuto1 not in range(0,60) or
            minuto2 not in range(0,60)
        ):
            raise(HorarioInvalidoException("Error: Los datos de horario deben estar en un rango de 0-23 horas y 0-59 minutos."))

        if time(hora1, minuto1) >= time(hora2, minuto2):
            raise(HorarioInvalidoException("La hora de cierre no puede ser menor que la de apertura"))
        
        indice = self.indice_horario(nombre_horario, dia)

        # Verificacion final para ver si modifica o agrega
        if indice==-1:
            self.agregar_horario(nombre_horario,dia,hora1,hora2, minuto1, minuto2, cantidad, edificio)
        else:
            self.horarios.at[indice, 'Día'] = dia
            self.horarios.at[indice, 'Hora de inicio'] = f"{hora1}:{minuto1}"
            self.horarios.at[indice, 'Hora de fin'] = f"{hora2}:{minuto2}"
            self.horarios.at[indice, 'Cantidad de alumnos'] = int(cantidad)
            self.horarios.at[indice, 'Edificio preferencial'] = edificio

    def eliminar_horario(self, nombre_materia:str, dia:str):
        """Metodo para eliminar una materia segun nombre y dia simultaneos."""
        condicion = (self.horarios['Nombre'] == nombre_materia) & (self.horarios['Día'] == dia)
        self.horarios = self.horarios[~condicion].reset_index(drop=True)


    def indice_horario(self, nombre_horario:str, dia:str):
        """Metodo usado como filtro para encontrar un dato de horario segun nombre y dia.
        Si no existe, retorna -1."""
        filtro = self.horarios[(self.horarios['Nombre'] == nombre_horario) & (self.horarios['Día'] == dia)]

        ret = -1
        try:
            ret = filtro.index[0]
        except IndexError as e:
            return ret


##### INTERFACE CON BACKEND

    def asignacion_automatica(self):



        aulas_backend = funciones_de_traduccion.traducir_aulas(self.aulas)
        horarios_backend = funciones_de_traduccion.traducir_clases(self.horarios)        
        asignar(horarios_backend, aulas_backend)


        # Sintaxis para el mapeo
        index_to_aula = self.aulas['Aula']
        index_to_edificio = self.aulas['Edificio']
        self.horarios['Aula asignada'] = horarios_backend['aula_asignada'].map(index_to_aula)
        self.horarios['Edificio asignado'] = horarios_backend['aula_asignada'].map(index_to_edificio)


    def exportar_horarios(self, path):

        self.horarios.to_excel(path, index=False)

        # Trabajar con el excel para que se termine viendo bonito:
        wb = load_workbook(path)
        ws = wb.active

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add padding
            ws.column_dimensions[col_letter].width = adjusted_width

        # Save changes
        wb.save(path)
        


