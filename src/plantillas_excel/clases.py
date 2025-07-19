'''
Este script crea la plantilla para el archivo excel de clases/horarios/aulas.

La plantilla está basada en el archivo excel que les directores de carreras le
pasan al encargado de la asignación. Se modificaron algunas cosas para que sea
más usable.

La plantilla tiene un preámbulo que consiste de:
- El logo de la universidad
- El nombre de la carrera
- El año y número de cuatrimestre

Abajo del preámbulo hay una tabla con las siguientes columnas:
- Año
- Materia
- Cuatrimestral / Anual
- Comisión
- Teórica / Práctica
- Cupo
- Día de la semana
- Horario de inicio
- Horario de fin
- Docente
- Auxiliar
- Promocionable (Si (nota) / No)
- Lugar
- Aula

La tabla tiene una fila de la tabla por cada clase, con celdas unidas
verticalmente (por ejemplo, la columna "Materia" está unida en las filas de las
clases de cada materia).
'''
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.styles.borders import Border
from openpyxl import Workbook

import sys

from estilos import (
    get_logo,
    fill_rojo_unrn,
    font_default,
    font_bold,
    font_preámbulo,
    centrado,
    a_la_derecha,
    a_la_izquierda,
    borde_blanco,
    borde_negro,
    todos_los_bordes_negros
)

from validadores import (
    no_cambiar_este_valor,
    año_del_plan_de_estudios,
    número_natural,
    día_de_la_semana,
    horario
)

COLUMNAS = (
    'Año',
    'Materia',
    'Cuatrimestral / Anual',
    'Comisión',
    'Teórica / Práctica',
    'Día',
    'Horario inicio',
    'Horario fin',
    'Cupo',
    'Docente',
    'Auxiliar',
    'Promocionable\nSi (Nota) / No',
    'Lugar',
    'Aula'
)

n_columns = len(COLUMNAS)
max_column = get_column_letter(n_columns)

def insertar_preámbulo(hoja: Worksheet):
    # Configurar tamaño de las filas
    altura_filas = font_preámbulo.size + 7
    hoja.row_dimensions[1].height = altura_filas
    hoja.row_dimensions[2].height = altura_filas

    # Marcar bordes externos
    hoja.column_dimensions['A'].border = Border(left=borde_negro)
    hoja.column_dimensions[max_column].border = Border(right=borde_negro)

    # Insertar el logo
    altura_logo = 2 * altura_filas - 2
    logo = get_logo(altura_logo)
    hoja.add_image(logo, 'A1')
    hoja.merge_cells(start_row=1, end_row=2, start_column=1, end_column=2)
    no_cambiar_este_valor.add(f'A1:B2')
    hoja.cell(1, 1).border = Border(left=borde_negro, top=borde_negro)
    hoja.cell(1, 2).border = Border(right=borde_negro, top=borde_negro)
    hoja.cell(2, 1).border = Border(left=borde_negro, bottom=borde_negro)
    hoja.cell(2, 2).border = Border(right=borde_negro, bottom=borde_negro)

    # Fila con el nombre de la carrera
    cell = hoja.cell(1, 3, value='Carrera: ')
    no_cambiar_este_valor.add(cell)
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_derecha
    cell.border = Border(top=borde_negro, bottom=borde_blanco, left=borde_negro)

    hoja.merge_cells(start_row=1, end_row=1, start_column=4, end_column=11)
    cell = hoja.cell(1, 4, value='') # Celda para completar el nombre de la carrera
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_izquierda
    
    hoja.merge_cells(start_row=1, end_row=1, start_column=12, end_column=n_columns)
    no_cambiar_este_valor.add(f'{get_column_letter(12)}1:{max_column}1')
    cell = hoja.cell(1, 12)
    cell.fill = fill_rojo_unrn

    for i in range(4, n_columns):
        hoja.cell(1, i).border = Border(top=borde_negro, bottom=borde_blanco)
    hoja.cell(1, n_columns).border = Border(top=borde_negro, bottom=borde_blanco, right=borde_negro)

    # Fila con el año y cuatrimestre
    cell = hoja.cell(2, 3, value='Año: ')
    no_cambiar_este_valor.add(cell)
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_derecha
    cell.border = Border(left=borde_negro)

    cell = hoja.cell(2, 4, value='') # Celda para completar el año
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_izquierda

    hoja.merge_cells(start_row=2, end_row=2, start_column=5, end_column=6)
    no_cambiar_este_valor.add(f'{get_column_letter(5)}2:{get_column_letter(6)}2')
    cell = hoja.cell(2, 5, value='Cuatrimestre: ')
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_derecha

    hoja.merge_cells(start_row=2, end_row=2, start_column=7, end_column=11)
    cell = hoja.cell(2, 7, value='') # Celda para completar el cuatrimestre
    cell.fill = fill_rojo_unrn
    cell.font = font_preámbulo
    cell.alignment = a_la_izquierda

    hoja.merge_cells(start_row=2, end_row=2, start_column=12, end_column=n_columns)
    no_cambiar_este_valor.add(f'{get_column_letter(12)}2:{max_column}2')
    hoja.cell(2, 12).fill = fill_rojo_unrn
    hoja.cell(2, n_columns).border = Border(right=borde_negro)

def insertar_tabla(hoja: Worksheet):
    # Intertar fila con los nombres de las columnas
    hoja.append(COLUMNAS)
    fila_header = hoja.max_row
    no_cambiar_este_valor.add(f'A{fila_header}:{max_column}{fila_header}')

    # Bloquear movimiento de los nombres para que se mantangas visibles al escrolear
    hoja.freeze_panes = hoja.cell(fila_header+1, 1)

    # Configurar estilo de los nombres
    for i in range(1, n_columns+1):
        cell = hoja.cell(fila_header, i)
        cell.font = font_bold
        cell.alignment = centrado
        cell.border = todos_los_bordes_negros
    
    # Ajustar tamaños de las columnas
    font_size_ratio = font_bold.size / 11
    hoja.column_dimensions['A'].width =  5 * font_size_ratio # Año
    hoja.column_dimensions['B'].width = 25 * font_size_ratio # Materia
    hoja.column_dimensions['C'].width = 12 * font_size_ratio # Cuatrimestral o anual
    hoja.column_dimensions['D'].width = 10 * font_size_ratio # Comisión
    hoja.column_dimensions['E'].width = 10 * font_size_ratio # Teórica o práctica
    hoja.column_dimensions['F'].width = 11 * font_size_ratio # Día
    hoja.column_dimensions['G'].width =  7 * font_size_ratio # Horario de inicio
    hoja.column_dimensions['H'].width =  7 * font_size_ratio # Horario de fin
    hoja.column_dimensions['I'].width =  5 * font_size_ratio # Cupo
    hoja.column_dimensions['J'].width = 12 * font_size_ratio # Docente
    hoja.column_dimensions['K'].width = 12 * font_size_ratio # Auxiliar
    hoja.column_dimensions['L'].width = 14 * font_size_ratio # Promocionable
    hoja.column_dimensions['M'].width = 12 * font_size_ratio # Lugar
    hoja.column_dimensions['N'].width =  8 * font_size_ratio # Aula

    # Agregar validadores
    año_del_plan_de_estudios.add(f'A{fila_header+1}:A1048576') # 1048576 significa hasta el final de la columna.
    día_de_la_semana.add(f'F{fila_header+1}:F1048576') # Día
    horario.add(f'G{fila_header+1}:G1048576') # Horario de inicio
    horario.add(f'H{fila_header+1}:H1048576') # Horario de fin
    número_natural.add(f'I{fila_header+1}:I1048576') # Cupo 

def crear_plantilla() -> Workbook:
    plantilla = Workbook()
    plantilla._fonts[0] = font_default

    hoja = plantilla.active
    hoja.title = 'Plantilla'
    hoja.add_data_validation(no_cambiar_este_valor)
    hoja.add_data_validation(año_del_plan_de_estudios)
    hoja.add_data_validation(número_natural)
    hoja.add_data_validation(día_de_la_semana)
    hoja.add_data_validation(horario)
    
    insertar_preámbulo(hoja)
    insertar_tabla(hoja)

    return plantilla

def main():
    plantilla = crear_plantilla()
    plantilla.save(sys.argv[1])

if __name__ == '__main__':
    main()
